import os
import sys
import argparse
import datetime
import shutil
import subprocess
import re
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side


class Result(object):

    def __init__(self, havedb=False):
        self.havedb = havedb
        self.border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000'))
        if self.havedb:
            from todb import ToDB
            self.db = ToDB()

    def get_log_list(self, path):
        result = []
        logs = subprocess.check_output('ls {}/*.txt'.format(path), shell=True)
        logs = logs.split('\n')
        del logs[-1]
    
        for log in logs:
            log = re.match('{}/(.+)\.txt'.format(path), log).group(1)
            result.append(log)
    
        return result
    
    def create_sheets(self, logs, wb):
        config_types = []
        for log in logs:
            #rbd_randrw_4k_runtime30_iodepth1_numjob1_imagenum2_hahaha_%100_2017_07_18_17_27_04
            configs_log = log.split('_')
            for configs in configs_log:
                if re.match('pool', configs):
                    pool_log = configs
                elif re.match('\d+[kM]', configs):
                    bs_log = configs
                elif re.match('%', configs):
                    rpercent_log = configs
                elif re.match('runtime', configs):
                    runtime_log = configs
                elif re.match('rw|randrw', configs):
                    rw_log = configs
                elif re.match('imagenum', configs):
                    imagenum_log = configs
                elif re.match('iodepth', configs):
                    iodepth_log = configs
                elif re.match('numjob', configs):
                    numjob_log = configs

            config_type = '{}_{}{}'.format(bs_log, rpercent_log, rw_log)
            if config_types.count(config_type) == 0:
                config_types.append(config_type)
        print config_types
        for config_type in config_types:
            ws = wb.create_sheet()
            ws.title = config_type
    
        wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
        self.fill_first_line(wb)
        return config_types
    
    def fill_first_line(self, wb):
        first_line = ['casename', 'case', 'blocksize', 'iodepth', 'numberjob', 'imagenum/client', 'read iops', 'write iops', 'total iops', 'read_write', 'latency(ms)', 'read bw(MB/s)', 'write bw(MB/s)', 'total bw(MB/s)', 'iops_community', 'latency(ms)_community', 'IOPS compare', 'Latency compare']
        for ws in wb:
            for i in range(len(first_line)):
                c = ws.cell(row = 1, column = (i+1))
                c.value = first_line[i]
                c.border = self.border
                c.font = Font(size=9)
    
            ws.column_dimensions["A"].width = 50.0
            ws.column_dimensions["B"].width = 10.0
    
    def get_bs(self, result, bs_log):
        match_nuit = re.match('(.*\d)(\w+)', result)
    
        if match_nuit.group(2) == 'B':
            result = '{}k'.format(int(match_nuit.group(1)) / 1024)
        elif match_nuit.group(2) == 'KiB':
            result = '{}k'.format(int(float(match_nuit.group(1))))
        elif match_nuit.group(2) == 'MiB':
            result = '{}k'.format(int(float(match_nuit.group(1))) * 1024)
        else:
            raise Exception("Error: Unrecognized block size unit {} in log file.".format(match_nuit.group(2)))
        match_log = re.search('(\d+)M', bs_log)
        if match_log:
            bs_log = '{}k'.format(int(match_log.group(1)) * 1024)

        if result == bs_log:
            return result
        else:
            raise Exception("Error: block size in log does not match log file name!")

    def fill_bs(self, bs_log, ws, row, column):
        ws.cell(row = row, column = column).value = bs_log
        ws.cell(row = row, column = column).border = self.border
        ws.cell(row = row, column = column).number_format = 'General'

    def get_iodepth(self, result, iodepth_log):
        match_iodepth = re.match('iodepth(\d+)', iodepth_log)
        if result == match_iodepth.group(1):
            return result
        else:
            raise Exception("Error: iodepth in log does not match log file name!")
 
    def fill_iodepth(self, result, ws, row, column):
        ws.cell(row = row, column = column).value = int(result)
        ws.cell(row = row, column = column).border = self.border
        ws.cell(row = row, column = column).number_format = 'General'

    def get_numjob(self, result, numjob_log):
        match_numjob = re.match('numjob(\d+)', numjob_log)
        if result == match_numjob.group(1):
            return result
        else:
            raise Exception("Error: numberjob in log does not match log file name!")
 
    def fill_numjob(self, result, ws, row, column):
        ws.cell(row = row, column = column).value = int(result)
        ws.cell(row = row, column = column).border = self.border
        ws.cell(row = row, column = column).number_format = 'General'

    def get_imagenum(self, path, results, imagenum_log):

        match_imagenum = re.match('imagenum(\d+)', imagenum_log)
        with open('{}/../fioserver_list.conf'.format(path), 'r') as f:
            clients = f.readlines()
            num_clients = len(clients)
    
        if len(results) == int(match_imagenum.group(1)) * num_clients:
            return match_imagenum.group(1), num_clients
        else:
            raise Exception("Error: image number in log does not match log file name!")

    def fill_imagenum(self, imagenum, clientnum, ws, row, column):
        ws.cell(row = row, column = column).value = '{}/{}'.format(imagenum, clientnum)
        ws.cell(row = row, column = column).border = self.border

    def get_readwrite(self, result, rw_log, rpercent_log):
        if result.lower() == rw_log.lower():
            return rw_log, rpercent_log
        else:
            raise Exception("Error: rw in log does not match log file name!")
        rpercent_log = re.sub('%', '', rpercent_log)

    def fill_readwrite(self, rw_log, rpercent_log, ws, row, column):
        ws.cell(row = row, column = column).value = '{}{}'.format(rw_log, rpercent_log)
        ws.cell(row = row, column = column).border = self.border
    
    def get_iops(self, log_list):
        r_iops = 0
        w_iops = 0
        for i in range(len(log_list)):
            if re.match(r'   read:', log_list[i]):
                match = re.match(r'\s*read: IOPS=([^k]+)(k?),', log_list[i])
                if re.search('k', match.group(2)):
                    r_iops = int(float(match.group(1)) * 1000)
                else:
                    r_iops = int(match.group(1))
            elif re.match(r'  write:', log_list[i]):
                match = re.match(r'\s*write: IOPS=([^k]+)(k?),', log_list[i])
                if re.search('k', match.group(2)):
                    w_iops = int(float(match.group(1)) * 1000)
                else:
                    w_iops = int(match.group(1))
        return r_iops, w_iops

    def fill_iops(self, r_iops, w_iops, ws, row, column):
        iops = r_iops + w_iops
        ws.cell(row = row, column = column).value = r_iops
        ws.cell(row = row, column = column).border = self.border
        ws.cell(row = row, column = column + 1).value = w_iops
        ws.cell(row = row, column = column + 1).border = self.border
        ws.cell(row = row, column = column + 2).value = iops
        ws.cell(row = row, column = column + 2).border = self.border

    def get_bw(self, log_list):
        read_bw = 0
        write_bw = 0
        for i in range(len(log_list)):
            if re.match(r'   read:', log_list[i]):
                match = re.match(r'\s*read: IOPS=.+k?, BW=(\d+\.?\d*)(\S+) ', log_list[i])
                unit = match.group(2)
                if unit == 'B/s':
                    read_bw = float(match.group(1)) / 1000000
                elif re.match(r'Ki', unit):
                    read_bw = float(match.group(1)) / 1000
                elif re.match(r'Mi', unit):
                    read_bw = float(match.group(1))
                else:
                    raise Exception("Error: Unrecognized BW unit {}.".format(unit))
            elif re.match(r'  write:', log_list[i]):
                match = re.match(r'\s*write: IOPS=.+k?, BW=(\d+\.?\d*)(\S+)', log_list[i])
                unit = match.group(2)
                if unit == 'B/s':
                    read_bw = float(match.group(1)) / 1000000
                elif re.match(r'Ki', unit):
                    write_bw = float(match.group(1)) / 1000
                elif re.match(r'Mi', unit):
                    write_bw = float(match.group(1))
                else:
                    raise Exception("Error: Unrecognized BW unit {}.".format(match.group(2)))
    
        return read_bw, write_bw

    def fill_bw(self, read_bw, write_bw, ws, row, column):
        bw = read_bw + write_bw
        ws.cell(row = row, column = column).value = read_bw
        ws.cell(row = row, column = column).border = self.border
        ws.cell(row = row, column = column + 1).value = write_bw
        ws.cell(row = row, column = column + 1).border = self.border
        ws.cell(row = row, column = column + 2).value = bw
        ws.cell(row = row, column = column + 2).border = self.border

    def get_lat(self, log_list):
        lat = 0
        for i in range(len(log_list)):
            match_lat = re.match(r'     lat \((.*)\):.*avg=(.*),', log_list[i])
            _lat = 0
            if match_lat:
                if match_lat.group(1) == 'msec':
                    _lat = float(match_lat.group(2))
                elif match_lat.group(1) == 'usec':
                    _lat = float(match_lat.group(2)) / 1000
                else:
                    raise Exception("Error: Unrecognized lat unit {}.".format(match_lat.group(1)))
            if _lat > lat:
                lat = _lat    
        return lat

    def fill_lat(self, lat, ws, row, column):
        ws.cell(row = row, column = column).value = lat
        ws.cell(row = row, column = column).border = self.border

    def get_job_info(self, log_dir):
        pwd_dir = log_dir.split('/')
        jobtime_match = re.search('_(\d{4})_(\d{2})_(\d{2})_(\d{2})_(\d{2})_(\d{2})', pwd_dir[-1])
        job_start_time = '{}-{}-{} {}:{}:{}'.format(
            jobtime_match.group(1),
            jobtime_match.group(2),
            jobtime_match.group(3),
            jobtime_match.group(4),
            jobtime_match.group(5),
            jobtime_match.group(6),
        )

        pwd_dir[-1] = re.sub('_(\d{4})_(\d{2})_(\d{2})_(\d{2})_(\d{2})_(\d{2})', '', pwd_dir[-1])

        pwd_log_dir = pwd_dir[-1].split('_')
        jobname = pwd_log_dir[0]
        del pwd_log_dir[0]

        cephconfig = ''
        for c in pwd_log_dir:
            cephconfig = cephconfig + c
        if cephconfig == '':
            cephconfig = 'Default'

        return job_start_time, jobname

    
    def fill_data_v2_21(self, log_dir, logs, wb, sheet_list):
        row_dic = {}
        for sheet in sheet_list:
            row_dic[sheet] = 2
        for log in logs:
            #rbd_randrw_4k_runtime30_iodepth1_numjob1_imagenum2_hahaha_%100_2017_07_18_17_27_04
            log_list = []
            with open('{}/{}.txt'.format(log_dir, log), 'r') as f:
                begin_to = False
                lines = f.readlines()
                case_status = lines[-1].strip()
                for line in lines:
                    if begin_to:
                        log_list.append(line)
                    if re.match('All clients', line):
                        begin_to = True

            configs_log = log.split('_')
            for configs in configs_log:
                if re.match('case\d+', configs):
                    casenum_log = configs
                elif re.match('pool', configs):
                    pool_log = configs
                elif re.match('\d+[kM]', configs):
                    bs_log = configs
                elif re.match('%', configs):
                    rpercent_log = configs
                elif re.match('runtime', configs):
                    runtime_log = configs
                elif re.match('rw|randrw', configs):
                    rw_log = configs
                elif re.match('imagenum', configs):
                    imagenum_log = configs
                elif re.match('iodepth', configs):
                    iodepth_log = configs
                elif re.match('numjob', configs):
                    numjob_log = configs

            bs = bs_log
            readwrite = '{}{}'.format(rw_log, rpercent_log)
            imagenum = re.sub('imagenum', '', imagenum_log)
            iodepth = re.sub('iodepth', '', iodepth_log)
            numjob = re.sub('numjob', '', numjob_log)
            r_iops = 0
            w_iops = 0
            iops = 0
            lat = 0
            r_bw = 0
            w_bw = 0
            bw = 0
            with open('{}/../fioserver_list.conf'.format(log_dir), 'r') as f:
                clients = f.readlines()
                clientnum = len(clients)
           
            if case_status == "Pass":
                config_type = '{}_{}{}'.format(bs_log, rpercent_log, rw_log)
                ws = wb.get_sheet_by_name(config_type)
                row = row_dic[config_type]
                row_dic[config_type] = row_dic[config_type] +1
        
                #fill imagenum, readwrite, bs, iodepth
                try:
                    results = subprocess.check_output('grep iodepth {}/{}.txt'.format(log_dir, log), shell=True).split('\n')
                    del results[-1]
                    result_match = re.search(r'rbd_image\d+:.*rw=(.*), bs=\(R\) (.*)-.*-.*-.*, ioengine=(.*), iodepth=(.*)', results[0])
                except Exception, e:
                    pass
                else:
                    imagenum, clientnum = self.get_imagenum(log_dir, results, imagenum_log)
                    self.fill_imagenum(imagenum, clientnum, ws, row, 6)

                    rw, rpercent = self.get_readwrite(result_match.group(1), rw_log, rpercent_log)
                    self.fill_readwrite(rw, rpercent, ws, row, 10)

                    bs = self.get_bs(result_match.group(2), bs_log)
                    self.fill_bs(bs, ws, row, 3)

                    if result_match.group(3) != 'rbd':
                        raise Exception("Error: The ioengine in log is not 'rbd'!")
                    iodepth = self.get_iodepth(result_match.group(4), iodepth_log)
                    self.fill_iodepth(iodepth, ws, row, 4)
        
                #fill numberjob
                try:
                    results = subprocess.check_output('grep jobs= {}/{}.txt'.format(log_dir, log), shell=True).split('\n')
                    result_match = re.search(r'jobs=(\d+)\)', results[0])
                except Exception, e:
                    pass
                else:
                    numjob = self.get_numjob(result_match.group(1), numjob_log)
                    self.fill_numjob(numjob, ws, row, 5)
        
                ws.cell(row = row, column = 2).value = '{}_{}_{}_{}'.format(bs, iodepth, numjob, imagenum)
                ws.cell(row = row, column = 2).border = self.border
                ws.cell(row = row, column = 1).value = '{}'.format(log)
                ws.cell(row = row, column = 1).border = self.border
                #fill iops and lat
                if len(log_list) == 0:
                    with open('{}/{}.txt'.format(log_dir, log), 'r') as f:
                        log_list = f.readlines()
                r_iops, w_iops = self.get_iops(log_list)
                self.fill_iops(r_iops, w_iops, ws, row, 7)
                lat = self.get_lat(log_list)
                self.fill_lat(lat, ws, row, 11)
                r_bw, w_bw = self.get_bw(log_list)
                self.fill_bw(r_bw, w_bw, ws, row, 12)

    def deal_with_fio_data(self, log_dir, log, jobid):
        job_start_time, jobname = self.get_job_info(log_dir)
        log_list = []
        with open('{}/{}.txt'.format(log_dir, log), 'r') as f:
            begin_to = False
            lines = f.readlines()
            case_status = lines[-1].strip()
            for line in lines:
                if begin_to:
                    log_list.append(line)
                if re.match('All clients', line):
                    begin_to = True
        time_match = re.search('_(\d{4})_(\d{2})_(\d{2})_(\d{2})_(\d{2})_(\d{2})', log)

        time = '{}-{}-{} {}:{}:{}'.format(
            time_match.group(1),
            time_match.group(2),
            time_match.group(3),
            time_match.group(4),
            time_match.group(5),
            time_match.group(6),
        )

        configs_log = log.split('_')
        for configs in configs_log:
            if re.match('case\d+', configs):
                casenum_log = configs
            elif re.match('pool', configs):
                pool_log = configs
            elif re.match('\d+[kM]', configs):
                bs_log = configs
            elif re.match('%', configs):
                rpercent_log = configs
            elif re.match('runtime', configs):
                runtime_log = configs
            elif re.match('rw|randrw', configs):
                rw_log = configs
            elif re.match('imagenum', configs):
                imagenum_log = configs
            elif re.match('iodepth', configs):
                iodepth_log = configs
            elif re.match('numjob', configs):
                numjob_log = configs
        
        configs_log.remove(casenum_log)
        configs_log.remove(pool_log)
        configs_log.remove(bs_log)
        configs_log.remove(rpercent_log)
        configs_log.remove(runtime_log)
        configs_log.remove(rw_log)
        configs_log.remove(imagenum_log)
        configs_log.remove(iodepth_log)
        configs_log.remove(numjob_log)
        configs_log.remove(jobname)
        configs_log.remove(time_match.group(1))
        configs_log.remove(time_match.group(2))
        configs_log.remove(time_match.group(3))
        configs_log.remove(time_match.group(4))
        configs_log.remove(time_match.group(5))
        configs_log.remove(time_match.group(6))

        print configs_log

        bs = bs_log
        readwrite = '{}{}'.format(rw_log, rpercent_log)
        imagenum = re.sub('imagenum', '', imagenum_log)
        iodepth = re.sub('iodepth', '', iodepth_log)
        numjob = re.sub('numjob', '', numjob_log)
        r_iops = 0
        w_iops = 0
        iops = 0
        lat = 0
        r_bw = 0
        w_bw = 0
        bw = 0
        with open('{}/../fioserver_list.conf'.format(log_dir), 'r') as f:
            clients = f.readlines()
            clientnum = len(clients)

        if case_status == "Pass":
            try:
                results = subprocess.check_output('grep iodepth {}/{}.txt'.format(log_dir, log), shell=True).split('\n')
                del results[-1]
                result_match = re.search(r'rbd_image\d+:.*rw=(.*), bs=\(R\) (.*)-.*-.*-.*, ioengine=(.*), iodepth=(.*)', results[0])
            except Exception, e:
                case_status = "Fail"
            else:
                imagenum, clientnum = self.get_imagenum(log_dir, results, imagenum_log)
                rw, rpercent = self.get_readwrite(result_match.group(1), rw_log, rpercent_log)
                readwrite = rw+rpercent
                bs = self.get_bs(result_match.group(2), bs_log)
                if result_match.group(3) != 'rbd':
                    raise Exception("Error: The ioengine in log is not 'rbd'!")
                iodepth = self.get_iodepth(result_match.group(4), iodepth_log)
    
            try:
                results = subprocess.check_output('grep jobs= {}/{}.txt'.format(log_dir, log), shell=True).split('\n')
                result_match = re.search(r'jobs=(\d+)\)', results[0])
            except Exception, e:
                case_status = "Fail"
            else:
                numjob = self.get_numjob(result_match.group(1), numjob_log)
    
            if len(log_list) == 0:
                with open('{}/{}.txt'.format(log_dir, log), 'r') as f:
                    log_list = f.readlines()
            r_iops, w_iops = self.get_iops(log_list)
            iops = r_iops + w_iops
            lat = self.get_lat(log_list)
            r_bw, w_bw = self.get_bw(log_list)
            bw = r_bw + w_bw

        if self.havedb:
            result_to_db = {
                'jobid': jobid,
                'time': time,
                'status': case_status,
                'case_name': log,
                'blocksize': bs,
                'iodepth': iodepth,
                'numberjob': numjob,
                'imagenum': imagenum,
                'clientnum': clientnum,
                'r_iops': r_iops,
                'w_iops': w_iops,
                'iops': iops,
                'readwrite': readwrite,
                'lat': lat,
                'r_bw': r_bw,
                'w_bw': w_bw,
                'bw': bw,
            }
            self.db.insert_tb_result(**result_to_db)
        return case_status

    def deal_with_fio_data_toexcel(self, file_name, log_dir):
        logs = self.get_log_list(log_dir)

        result_table = Workbook()
        sheet_list = self.create_sheets(logs, result_table)
        self.fill_data_v2_21(log_dir, logs, result_table, sheet_list)

        result_table.save('{}/{}.xlsx'.format(log_dir, file_name))
        return '{}/{}.xlsx'.format(log_dir, file_name)


def main():
    parser = argparse.ArgumentParser(
        prog="result",
        version='v0.1',
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
        description="generate excel file from fio log")
    parser.add_argument('-N', '--suitename', dest="suitename",
                metavar="test suite name", action="store",
                    help='''test suite name''')
    parser.add_argument('-T', '--timetag', dest="timetag",
                metavar="test suite test time", action="store",
                    help='''test suite test time''')
    args = parser.parse_args()

    result = Result()
    result.deal_with_fio_data(args.suitename, args.timetag, 'result_{}'.format(args.timetag))

if __name__ == '__main__':
    main()
