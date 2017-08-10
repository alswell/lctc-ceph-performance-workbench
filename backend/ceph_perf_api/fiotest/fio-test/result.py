import os
import sys
import argparse
import datetime
import shutil
import subprocess
import re
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side

from collect_system_status import SysInfo


class Result(object):

    def __init__(self, havedb=False):
        self.havedb = havedb
        self.border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000'))
        if self.havedb:
            from todb import ToDB
            self.db = ToDB()

    def get_log_list(self, suitename, tag):
        result = []
        pwd_path = os.getcwd().split('/')
        if pwd_path[-1] == suitename and pwd_path[-2] == 'test-suites':
            path = "{}/{}".format(os.getcwd(), tag)
        else:
            path = "{}/test-suites/{}/{}".format(os.getcwd(), suitename, tag)
        os.chdir(path)
        logs = subprocess.check_output('ls {}/*.txt'.format(path), shell=True)
        logs = logs.split('\n')
        del logs[-1]
    
        for log in logs:
            log = re.match('{}/(.+)\.txt'.format(path), log).group(1)
            result.append(log)
    
        return path, result
    
    def create_sheets(self, logs, wb):
        config_types = []
        for log in logs:
            #rbd_randrw_4k_runtime30_iodepth1_numjob1_imagenum2_hahaha_%100_2017_07_18_17_27_04
            config_list = log.split('_')
            config_type = '{}_{}{}'.format(config_list[2], config_list[8], config_list[1])
            if config_types.count(config_type) == 0:
                config_types.append(config_type)
        print config_types
        for config_type in config_types:
            ws = wb.create_sheet()
            ws.title = config_type
    
        wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
        self.fill_first_line(wb)
        return config_types
    
    def fill_first_line(self, wb):
        first_line = ['casename', 'case', 'blocksize', 'iodepth', 'numberjob', 'imagenum/client', 'iops', 'read_write', 'latency(ms)', 'bw(MB/s)', 'iops_community', 'latency(ms)_community', 'IOPS compare', 'Latency compare']
        for ws in wb:
            for i in range(len(first_line)):
                c = ws.cell(row = 1, column = (i+1))
                c.value = first_line[i]
                c.border = self.border
                c.font = Font(size=9)
    
            ws.column_dimensions["A"].width = 50.0
            ws.column_dimensions["B"].width = 15.0
    
    def fill_bs_C(self, result, config_log, ws, row):
        match_nuit = re.match('(.*\d)(\w+)', result)
    
        if match_nuit.group(2) == 'B':
            result = '{}k'.format(int(match_nuit.group(1)) / 1024)
        elif match_nuit.group(2) == 'KiB':
            result = '{}k'.format(int(float(match_nuit.group(1))))
    
        if result == config_log[2]:
            ws.cell(row = row, column = 3).value = config_log[2]
            ws.cell(row = row, column = 3).border = self.border
            ws.cell(row = row, column = 3).number_format = 'General'
        else:
            print "Error: block size in log does not match log file name!"
            sys.exit(1)
        return result
    
    def fill_iodepth_D(self, result, config_log, ws, row):
        match_iodepth = re.match('iodepth(\d+)', config_log[4])
        if result == match_iodepth.group(1):
            ws.cell(row = row, column = 4).value = int(result)
            ws.cell(row = row, column = 4).border = self.border
            ws.cell(row = row, column = 4).number_format = 'General'
        else:
            print "Error: iodepth in log does not match log file name!"
            sys.exit(1)
        return result
    
    def fill_numjob_E(self, result, config_log, ws, row):
        match_numjob = re.match('numjob(\d+)', config_log[5])
        if result == match_numjob.group(1):
            ws.cell(row = row, column = 5).value = int(result)
            ws.cell(row = row, column = 5).border = self.border
            ws.cell(row = row, column = 5).number_format = 'General'
        else:
            print "Error: numberjob in log does not match log file name!"
            sys.exit(1)
        return result
    
    def fill_imagenum_F(self, results, config_log, ws, row):
        match_imagenum = re.match('imagenum(\d+)', config_log[6])
        with open('../fioserver_list.conf', 'r') as f:
            clients = f.readlines()
            num_clients = len(clients)
    
        if len(results) == int(match_imagenum.group(1)) * num_clients:
            ws.cell(row = row, column = 6).value = '{}/{}'.format(match_imagenum.group(1), num_clients)
            ws.cell(row = row, column = 6).border = self.border
        else:
            print "Error: image number in log does not match log file name!"
            sys.exit(1)
        return match_imagenum.group(1), num_clients
    
    def fill_readwrite_H(self, result, config_log, ws, row):
        if result.lower() == config_log[1].lower():
            ws.cell(row = row, column = 8).value = '{}{}'.format(config_log[1], config_log[8])
            ws.cell(row = row, column = 8).border = self.border
        else:
            print "Error: rw in log does not match log file name!"
            sys.exit(1)
        return config_log[1]+config_log[8]
    
    def fill_iops_G(self, log_list, ws, row):
        read_iops = 0
        write_iops = 0
        for i in range(len(log_list)):
            if re.match(r'   read:', log_list[i]):
                match = re.match(r'\s*read: IOPS=(\d+),', log_list[i])
                read_iops = int(match.group(1))
            elif re.match(r'  write:', log_list[i]):
                match = re.match(r'\s*write: IOPS=(\d+),', log_list[i])
                write_iops = int(match.group(1))
    
        iops = read_iops + write_iops
        ws.cell(row = row, column = 7).value = iops
        ws.cell(row = row, column = 7).border = self.border
        return iops

    def fill_bw_J(self, log_list, ws, row):
        read_bw = 0
        write_bw = 0
        for i in range(len(log_list)):
            if re.match(r'   read:', log_list[i]):
                match = re.match(r'\s*read: IOPS=\d+, BW=(\d+)(\S+) ', log_list[i])
                if match.group(2) == 'B/s':
                    read_bw = float(match.group(1)) / 1000000
                elif match.group(2) == 'KiB/s':
                    read_bw = float(match.group(1)) / 1000
                elif match.group(2) == 'MiB/s':
                    read_bw = float(match.group(1))
                else:
                    print "Error: Unrecognized BW unit {}.".format(match.group(2))
                    sys.exit(1)
            elif re.match(r'  write:', log_list[i]):
                match = re.match(r'\s*write: IOPS=\d+, BW=(\d+)(\S+)', log_list[i])
                if match.group(2) == 'KiB/s':
                    write_bw = float(match.group(1)) / 1000
                elif match.group(2) == 'MiB/s':
                    write_bw = float(match.group(1))
                else:
                    print "Error: Unrecognized BW unit {}.".format(match.group(2))
                    sys.exit(1)
    
        bw = read_bw + write_bw
        ws.cell(row = row, column = 10).value = bw
        ws.cell(row = row, column = 10).border = self.border
        return bw
    
    def fill_lat_I(self, log_list, ws, row):
        lat = 0
        for i in range(len(log_list)):
            match_lat = re.match(r'     lat \((.*)\):.*avg=(.*),', log_list[i])
            _lat = 0
            if match_lat:
                if match_lat.group(1) == 'msec':
                    _lat = float(match_lat.group(2))
                elif match_lat.group(1) == 'usec':
                    _lat = float(match_lat.group(2)) / 1000
            if _lat > lat:
                lat = _lat
    
        ws.cell(row = row, column = 9).value = lat
        ws.cell(row = row, column = 9).border = self.border
        return lat
    
    
    def fill_data_v2_21(self, logs, wb, sheet_list):
        row_dic = {}
        for sheet in sheet_list:
            row_dic[sheet] = 2

        pwd_dir = os.getcwd().split('/')
        pwd_log_dir = pwd_dir[-1].split('_')
        job_start_time = '{}-{}-{} {}:{}:{}'.format(
            pwd_log_dir[-6],
            pwd_log_dir[-5],
            pwd_log_dir[-4],
            pwd_log_dir[-3],
            pwd_log_dir[-2],
            pwd_log_dir[-1],
        )
        for i in range(6):
            del pwd_log_dir[-1]
        jobname = pwd_log_dir[0]
        del pwd_log_dir[0]

        cephconfig = ''
        for c in pwd_log_dir:
            cephconfig = cephconfig + c
        if cephconfig == '':
            cephconfig = 'Default'

        for log in logs:
            #rbd_randrw_4k_runtime30_iodepth1_numjob1_imagenum2_hahaha_%100_2017_07_18_17_27_04
            config_log = log.split('_')
            config_type = '{}_{}{}'.format(config_log[2], config_log[8], config_log[1])
            ws = wb.get_sheet_by_name(config_type)
            row = row_dic[config_type]
            row_dic[config_type] = row_dic[config_type] +1
    
            #fill imagenum, readwrite, bs, iodepth
            results = subprocess.check_output('grep iodepth {}.txt'.format(log), shell=True).split('\n')
            del results[-1]
            result_match = re.search(r'rbd_image\d+:.*rw=(.*), bs=\(R\) (.*)-.*-.*-.*, ioengine=(.*), iodepth=(.*)', results[0])
    
            imagenum, clientnum = self.fill_imagenum_F(results, config_log, ws, row)
            readwrite = self.fill_readwrite_H(result_match.group(1), config_log, ws, row)
            bs = self.fill_bs_C(result_match.group(2), config_log, ws, row)
            if result_match.group(3) != 'rbd':
                print "Error: The ioengine in log is not 'rbd'!"
                sys.exit(1)
            iodepth = self.fill_iodepth_D(result_match.group(4), config_log, ws, row)
    
            #fill numberjob    
            results = subprocess.check_output('grep jobs= {}.txt'.format(log), shell=True).split('\n')
            result_match = re.search(r'jobs=(\d+)\)', results[0])
    
            numjob = self.fill_numjob_E(result_match.group(1), config_log, ws, row)
    
            ws.cell(row = row, column = 2).value = '{}_{}_{}_{}'.format(bs, iodepth, numjob, imagenum)
            ws.cell(row = row, column = 2).border = self.border
            ws.cell(row = row, column = 1).value = '{}'.format(log)
            ws.cell(row = row, column = 1).border = self.border
            #fill iops and lat
            log_list = []
            with open('{}.txt'.format(log), 'r') as f:
                begin_to = False
                lines = f.readlines()
                case_status = lines[-1].strip()
                for line in lines:
                    if begin_to:
                        log_list.append(line)
                    if re.match('All clients', line):
                        begin_to = True
            if len(log_list) == 0:
                with open('{}.txt'.format(log), 'r') as f:
                    log_list = f.readlines()
            iops = self.fill_iops_G(log_list, ws, row)
            lat = self.fill_lat_I(log_list, ws, row)
            bw = self.fill_bw_J(log_list, ws, row)

            if self.havedb:
                time = '{}-{}-{} {}:{}:{}'.format(
                    config_log[9],
                    config_log[10],
                    config_log[11],
                    config_log[12],
                    config_log[13],
                    config_log[14]
                )
                result_to_db = {
                    'jobtime': job_start_time,
                    'ceph_config': cephconfig,
                    'time': time,
                    'status': case_status,
                    'case_name': log,
                    'blocksize': bs,
                    'iodepth': iodepth,
                    'numberjob': numjob,
                    'imagenum': imagenum,
                    'clientnum': clientnum,
                    'iops': iops,
                    'readwrite': readwrite,
                    'lat': lat,
                    'bw': bw
                }
                self.db.insert_tb_result(**result_to_db)
        if self.havedb:
            self.db.close_db()

    def deal_with_fio_data(self, suitename, jobtag, file_name):
        log_dir, logs = self.get_log_list(suitename, jobtag)

        result_table = Workbook()
        sheet_list = self.create_sheets(logs, result_table)
        self.fill_data_v2_21(logs, result_table, sheet_list)

        result_table.save('./{}.xlsx'.format(file_name))
        return '{}/{}.xlsx'.format(log_dir, file_name)


def main():
    parser = argparse.ArgumentParser(
        prog="result",
        version='v0.1',
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
        description="generate excel file from fio log")
    parser.add_argument('-N', '--suitename', dest="suitename",
                metavar="test suite name", action="store",
                    help='''test suite name''')
    parser.add_argument('-T', '--timetag', dest="timetag",
                metavar="test suite test time", action="store",
                    help='''test suite test time''')
    args = parser.parse_args()

    result = Result()
    result.deal_with_fio_data(args.suitename, args.timetag, 'result_{}'.format(args.timetag))

if __name__ == '__main__':
    main()
